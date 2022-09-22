from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from item import Product, StaffMember
from datetime import datetime


# wb = load_workbook(filename= 'test.xlsx')
# print(wb)
# ws1 = wb.active
# ws1.title = 'sheet'
# sheet = wb['sheet']

# sheet['C16'].value = 5
# print(sheet['C16'].value)
#wb.save(filename='test.xlsx')

#send a nested list.
#create a method that takes the the first list splits the nsn into a list 
#and returns a new nested list [[split NSN], name, batch amount, qty?]  
#second method write all the list details to the excel doc, also need to active the doc
#method to save using date time stamp and close the doc.

class Order:
    """Add class doc here."""       
    
    def __init__ (self):
        #File variables need changing if the excel template changes.
        #excel template is called order_form.py version 1.0
        self.NSN_CELL_LENGTH = 13
        self.start_col = 3 #using get_column_letter this represents 'C'
        self.start_row = 16
        self.DESCRIP = get_column_letter(16)
        self.UNIT  = get_column_letter(17)
        self.SIGNATURE = 'E42'
        self.NAME = 'M42'
        self.STAFF_NUM = 'M43'
        self.EMAIL = 'T42'
        self.DATE = 'T43' 
        
        #----------------------------------------------------------
        self.file_incriment = 1
        self.ws = ''
        self.wb = ''
        
    def split_id (self, prod_id:str) -> list[str]:
        """@arg takes a string prod_id. 
        @return a list with the nsn split into 13 strings
        @raises a ValueError if the @arg is less or more than cell length."""
        if len(prod_id) == self.NSN_CELL_LENGTH:
            return list(prod_id)  
        raise ValueError(f'{prod_id} is not the correct length')  
    
    def active_file(self) -> None:
        """Prepares the excel file."""
        self.wb = load_workbook(filename= f'test1.xlsx')
        self.ws = self.wb.active
        self.ws.title = 'order1'
        
    def check_order_amount(self, prod_list:list[Product]) -> bool:  
        """@returns a bool if there are more than 20 orders."""
        return len(prod_list) > 20      
        
    def add_order_data(self):
        """Adds no more than 20 orders to a worksheet."""
        pass
    
    def write_prod_line(self, prod_id:str) -> None:
        """@arg is the product_id as a string. Calls split id.
        Adds the listed id values to the order form."""
        new_prod = self.split_id(prod_id)
        index = 0
        for cell in range(self.start_col, self.start_col + self.NSN_CELL_LENGTH):
            col = get_column_letter(cell)
            self.ws[col + str(self.start_row)].value = new_prod[index]
            index += 1
        print('ID added to form!!')
        

    def write_order_to_file(self, prod_dict:dict[str,list[Product|StaffMember]]) -> None:
        """Writes all the orders into the excel file."""
        prod_list = prod_dict['products'] #list of all products
        staff_obj = prod_dict['staff'][0] #the staff object
        if not self.check_order_amount(prod_list):
            #activate file
            self.active_file()
            #add staff details
            self.ws[self.SIGNATURE].value = staff_obj.staff_name
            self.ws[self.NAME].value = staff_obj.staff_name
            self.ws[self.STAFF_NUM].value = staff_obj.staff_id
            self.ws[self.EMAIL].value = staff_obj.staff_email
            #add product details
            for rec in prod_list:
                self.write_prod_line(rec.prod_id)
                self.ws[self.DESCRIP+str(self.start_row)].value = rec.discript
                self.ws[self.UNIT+str(self.start_row)].value = rec.order_amount
                self.start_row += 1
            #save and close the excel file
            self.save_close_file(staff_obj.staff_name)           
            
    def save_close_file(self, staff_name:str) -> None:
        """Save the file with a new file name and updates the database."""        
        time = self.date_time_stamp()
        self.wb.save(filename=f'{time}_{staff_name}.xlsx')
        print('New file created')
        
    def date_time_stamp(self) -> str:
        """Creates a time stamp for the current time in the format year/months/days
        /hours/min/secs."""
        time = datetime.now()
        return time.strftime("%Y%m%d%H%M%S")


    


# prod_id = '7777555555555'
# test= Order()
# test.active_file()
# test.write_prod_line(prod_id)
# test.save_close_file()